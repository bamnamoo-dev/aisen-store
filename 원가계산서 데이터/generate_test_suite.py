import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

OUT_DIR = r"G:\내 드라이브\antigravity\asisen-store\원가계산서 데이터"
FONT_PATH = "C:/Windows/Fonts/malgun.ttf"

# ReportLab 한글 폰트 등록
pdfmetrics.registerFont(TTFont('Malgun', FONT_PATH))

CASES = [
    {
        "id": "CASE_1",
        "title": "방학 중 교실 도색 공사",
        "trade": "건축",
        "period": 15,
        "note": "30일 미만 단기공사 3대보험(건강,연금,요양) 부당 계상 케이스",
        "items": [
            ("직접재료비", 6500000, "산출내역서 기준"),
            ("간접재료비", 0, ""),
            ("재료비 소계", 6500000, ""),
            ("직접노무비", 5200000, "산출내역서 기준"),
            ("간접노무비", 754000, "직접노무비 × 14.5%"),
            ("노무비 소계", 5954000, ""),
            ("기계경비 (산출경비)", 0, ""),
            ("기타경비", 666900, "(재료비+노무비) × 5.35%"),
            ("산재보험료", 192400, "노무비 × 3.7%"),
            ("고용보험료", 50960, "노무비 × 0.98%"),
            ("국민건강보험료", 213200, "위법 계상 (30일 미만 0원이어야 함)"),
            ("국민연금보험료", 234000, "위법 계상 (30일 미만 0원이어야 함)"),
            ("노인장기요양보험", 27600, "위법 계상 (30일 미만 0원이어야 함)"),
            ("산업안전보건관리비", 342810, "대상액 × 2.93%"),
            ("퇴직공제부금비", 0, "1억 미만 제외"),
            ("환경보전비", 0, "실내도색 제외"),
            ("경비 소계", 1727870, ""),
            ("순원가", 14181870, ""),
            ("일반관리비", 850900, "순원가 × 6.0%"),
            ("이윤", 1327330, "(노무비+경비+일반관리비) × 15.0%"),
            ("총원가", 16360100, ""),
            ("부가가치세", 1636010, "총원가 × 10%"),
            ("총도급액 (합계)", 17996110, "10원 미만 절사"),
        ]
    },
    {
        "id": "CASE_2",
        "title": "본관 교실 방충망 교체공사",
        "trade": "건축",
        "period": 7,
        "note": "2천만원 미만 산안비 면제대상 과다 계상 케이스",
        "items": [
            ("직접재료비", 2800000, "산출내역서 기준"),
            ("간접재료비", 0, ""),
            ("재료비 소계", 2800000, ""),
            ("직접노무비", 1900000, "산출내역서 기준 (대상액 470만)"),
            ("간접노무비", 275500, "14.5%"),
            ("노무비 소계", 2175500, ""),
            ("기계경비 (산출경비)", 0, ""),
            ("기타경비", 266180, "5.35%"),
            ("산재보험료", 70300, "3.7%"),
            ("고용보험료", 18620, "0.98%"),
            ("국민건강보험료", 0, "30일 미만 제외"),
            ("국민연금보험료", 0, "30일 미만 제외"),
            ("노인장기요양보험", 0, "30일 미만 제외"),
            ("산업안전보건관리비", 250000, "과다 계상 (2천만원 미만 면제/기준초과)"),
            ("퇴직공제부금비", 0, ""),
            ("환경보전비", 0, ""),
            ("경비 소계", 605100, ""),
            ("순원가", 5580600, ""),
            ("일반관리비", 334830, "6.0%"),
            ("이윤", 467310, "15.0%"),
            ("총원가", 6382740, ""),
            ("부가가치세", 638274, "10%"),
            ("총도급액 (합계)", 7021010, "10원 미만 절사"),
        ]
    },
    {
        "id": "CASE_3",
        "title": "체육관 LED 조명 개선 전기공사",
        "trade": "전기",
        "period": 45,
        "note": "일반관리비 및 이윤 법정 상한선 초과 케이스",
        "items": [
            ("직접재료비", 18000000, "산출내역서 기준"),
            ("간접재료비", 0, ""),
            ("재료비 소계", 18000000, ""),
            ("직접노무비", 9500000, "산출내역서 기준"),
            ("간접노무비", 1092500, "11.5%"),
            ("노무비 소계", 10592500, ""),
            ("기계경비 (산출경비)", 0, ""),
            ("기타경비", 1429620, "5.0%"),
            ("산재보험료", 351500, "3.7%"),
            ("고용보험료", 93100, "0.98%"),
            ("국민건강보험료", 390450, "사후정산 PS (정상)"),
            ("국민연금보험료", 427500, "사후정산 PS (정상)"),
            ("노인장기요양보험", 50560, "12.95% (정상)"),
            ("산업안전보건관리비", 764500, "2.78% (정상)"),
            ("퇴직공제부금비", 0, "전기 3억 미만 제외"),
            ("환경보전비", 0, ""),
            ("경비 소계", 3507230, ""),
            ("순원가", 32099730, ""),
            ("일반관리비", 2450000, "초과 계상 7.63% (법정상한 6.0% 초과)"),
            ("이윤", 3200000, "초과 계상 19.33% (법정상한 15.0% 초과)"),
            ("총원가", 37749730, ""),
            ("부가가치세", 3774970, "10% 절사"),
            ("총도급액 (합계)", 41524700, "10원 미만 절사"),
        ]
    },
    {
        "id": "CASE_4",
        "title": "여름방학 학생 화장실 환경개선공사",
        "trade": "건축",
        "period": 60,
        "note": "2026 조달청 제비율 100% 모범 준수 올그린 케이스",
        "items": [
            ("직접재료비", 52000000, "산출내역서 기준"),
            ("간접재료비", 0, ""),
            ("재료비 소계", 52000000, ""),
            ("직접노무비", 35000000, "산출내역서 기준"),
            ("간접노무비", 5075000, "직접노무비 × 14.5%"),
            ("노무비 소계", 40075000, ""),
            ("기계경비 (산출경비)", 0, ""),
            ("기타경비", 4926010, "(재료비+노무비) × 5.35%"),
            ("산재보험료", 1295000, "노무비 × 3.7%"),
            ("고용보험료", 343000, "노무비 × 0.98%"),
            ("국민건강보험료", 1438500, "사후정산 PS 준수 (4.11%)"),
            ("국민연금보험료", 1575000, "사후정산 PS 준수 (4.50%)"),
            ("노인장기요양보험", 186280, "건강보험 × 12.95%"),
            ("산업안전보건관리비", 2549100, "대상액 × 2.93%"),
            ("퇴직공제부금비", 805000, "1억 이상 건축 준수 (2.3%)"),
            ("환경보전비", 260000, "재료비 × 0.5% 준수"),
            ("경비 소계", 13377890, ""),
            ("순원가", 105452890, ""),
            ("일반관리비", 6327170, "순원가 × 6.0% 준수"),
            ("이윤", 8967000, "(노무비+경비+일반관리비) × 15.0% 준수"),
            ("총원가", 120747060, ""),
            ("부가가치세", 12074700, "총원가 × 10% 절사"),
            ("총도급액 (합계)", 132821760, "10원 미만 절사"),
        ]
    },
    {
        "id": "CASE_5",
        "title": "긴급 옥상 누수 우레탄 방수공사",
        "trade": "건축",
        "period": 10,
        "note": "총액 맞추고 산재/고용 필수보험 0원 고의 누락 꼼수 케이스",
        "items": [
            ("직접재료비", 2200000, "산출내역서 기준"),
            ("간접재료비", 0, ""),
            ("재료비 소계", 2200000, ""),
            ("직접노무비", 1500000, "산출내역서 기준"),
            ("간접노무비", 217500, "14.5%"),
            ("노무비 소계", 1717500, ""),
            ("기계경비 (산출경비)", 0, ""),
            ("기타경비", 209580, "5.35%"),
            ("산재보험료", 0, "고의 누락 (법정 필수 비목 0원)"),
            ("고용보험료", 0, "고의 누락 (법정 필수 비목 0원)"),
            ("국민건강보험료", 0, "30일 미만 제외"),
            ("국민연금보험료", 0, "30일 미만 제외"),
            ("노인장기요양보험", 0, "30일 미만 제외"),
            ("산업안전보건관리비", 108410, "2.93%"),
            ("퇴직공제부금비", 0, ""),
            ("환경보전비", 0, ""),
            ("경비 소계", 317990, ""),
            ("순원가", 4235490, ""),
            ("일반관리비", 254120, "6.0%"),
            ("이윤", 343440, "15.0%"),
            ("총원가", 4833050, ""),
            ("부가가치세", 483300, "10% 절사"),
            ("총도급액 (합계)", 5316350, "10원 미만 절사"),
        ]
    }
]

def make_excel(case):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "원가계산서"

    # 헤더
    ws.merge_cells("A1:D1")
    ws["A1"] = f"【공사원가계산서】 {case['title']}"
    ws["A1"].font = Font(name="맑은 고딕", size=15, bold=True, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws["A2"] = f"■ 공종: {case['trade']}공사  |  ■ 공사기간: {case['period']}일  |  ■ 비고: {case['note']}"
    ws["A2"].font = Font(name="맑은 고딕", size=9, color="475569")
    ws.merge_cells("A2:D2")

    headers = ["비 목 (항목명)", "금 액 (원)", "구성비", "산출 공식 및 법적 산출근거"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 24

    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_idx = 5
    for item, amt, note in case["items"]:
        c1 = ws.cell(row=row_idx, column=1, value=item)
        c2 = ws.cell(row=row_idx, column=2, value=amt)
        c2.number_format = "#,##0"
        c3 = ws.cell(row=row_idx, column=3, value="")
        c4 = ws.cell(row=row_idx, column=4, value=note)

        is_subtotal = any(k in item for k in ["소계", "순원가", "총원가", "총도급액", "합계"])
        for c in [c1, c2, c3, c4]:
            c.border = border
            c.font = Font(name="맑은 고딕", size=9.5, bold=is_subtotal)
            if is_subtotal:
                c.fill = PatternFill(start_color="F1F5F9", fill_type="solid")

        c1.alignment = Alignment(horizontal="left", indent=1 if not is_subtotal else 0)
        c2.alignment = Alignment(horizontal="right")
        c4.alignment = Alignment(horizontal="left")
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 45

    path = os.path.join(OUT_DIR, f"{case['id']}_{case['title']}.xlsx")
    wb.save(path)
    print(f"[XLSX 생성] {path}")
    return path

def make_pdf(case):
    path = os.path.join(OUT_DIR, f"{case['id']}_{case['title']}.pdf")
    c = canvas.Canvas(path, pagesize=A4)
    width, height = A4

    c.setFont('Malgun', 16)
    c.drawString(40, height - 45, f"【공사원가계산서】 {case['title']}")

    c.setFont('Malgun', 9)
    c.setFillColorRGB(0.3, 0.3, 0.4)
    c.drawString(40, height - 65, f"■ 공종: {case['trade']}공사  |  ■ 공사기간: {case['period']}일  |  ■ 특징: {case['note']}")

    y = height - 90
    c.setFillColorRGB(0.12, 0.23, 0.54)
    c.rect(40, y - 5, width - 80, 20, fill=True, stroke=False)
    c.setFillColorRGB(1, 1, 1)
    c.setFont('Malgun', 9.5)
    c.drawString(50, y, "비 목 (항목명)")
    c.drawString(200, y, "금 액 (원)")
    c.drawString(320, y, "산출 공식 및 산출근거")

    y -= 22
    for item, amt, note in case["items"]:
        is_subtotal = any(k in item for k in ["소계", "순원가", "총원가", "총도급액", "합계"])
        if is_subtotal:
            c.setFillColorRGB(0.94, 0.96, 0.98)
            c.rect(40, y - 4, width - 80, 16, fill=True, stroke=False)
            c.setFillColorRGB(0.05, 0.1, 0.25)
            c.setFont('Malgun', 8.5)
        else:
            c.setFillColorRGB(0.1, 0.1, 0.1)
            c.setFont('Malgun', 8)

        c.drawString(50, y, item)
        c.drawRightString(280, y, f"{amt:,} 원")
        c.drawString(320, y, note[:35])
        y -= 16

    c.save()
    print(f"[PDF 생성] {path}")
    return path

def make_image(case):
    img_w, img_h = 1000, 1150
    img = Image.new("RGB", (img_w, img_h), "#FFFFFF")
    draw = ImageDraw.Draw(img)

    f_title = ImageFont.truetype(FONT_PATH, 24)
    f_sub = ImageFont.truetype(FONT_PATH, 13)
    f_th = ImageFont.truetype(FONT_PATH, 14)
    f_td = ImageFont.truetype(FONT_PATH, 13)
    f_bold = ImageFont.truetype(FONT_PATH, 13)

    # 헤더
    draw.rectangle([0, 0, img_w, 75], fill="#0F172A")
    draw.text((30, 15), f"공 사 원 가 계 산 서 ({case['trade']}공사)", fill="#FFFFFF", font=f_title)
    draw.text((30, 48), f"공사명: {case['title']}  |  공사기간: {case['period']}일  |  특징: {case['note']}", fill="#94A3B8", font=f_sub)

    # 테이블 헤더
    draw.rectangle([30, 90, img_w - 30, 125], fill="#1E293B")
    draw.text((45, 98), "비목 (항목명)", fill="#F8FAFC", font=f_th)
    draw.text((320, 98), "금 액 (원)", fill="#F8FAFC", font=f_th)
    draw.text((520, 98), "산출 근거 및 요율", fill="#F8FAFC", font=f_th)

    y = 135
    for item, amt, note in case["items"]:
        is_sub = any(k in item for k in ["소계", "순원가", "총원가", "총도급액", "합계"])
        if is_sub:
            draw.rectangle([30, y - 3, img_w - 30, y + 21], fill="#F1F5F9")
            color = "#0F172A"
            font = f_bold
        else:
            color = "#334155"
            font = f_td

        draw.text((45, y), item, fill=color, font=font)
        draw.text((320, y), f"{amt:,}", fill=color, font=font)
        draw.text((520, y), note[:35], fill="#64748B", font=font)
        draw.line([(30, y + 23), (img_w - 30, y + 23)], fill="#E2E8F0", width=1)
        y += 28

    path = os.path.join(OUT_DIR, f"{case['id']}_{case['title']}.png")
    img.save(path)
    print(f"[PNG 생성] {path}")
    return path

if __name__ == "__main__":
    print(f"=== 가상 원가계산서 5종 × 3가지 포맷(총 15개) 생성 시작 ===")
    for c in CASES:
        make_excel(c)
        make_pdf(c)
        make_image(c)
    print(f"=== 총 15개 파일 생성 완료! 저장 위치: {OUT_DIR} ===")
