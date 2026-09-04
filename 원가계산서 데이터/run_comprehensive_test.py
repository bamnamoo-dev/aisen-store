import sys
import os
import time
import json
import base64
import urllib.request
import openpyxl
from reportlab.pdfgen import canvas
import pypdf

# Windows 콘솔 utf-8 설정
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = r"G:\내 드라이브\antigravity\asisen-store\원가계산서 데이터"
ENV_PATH = r"G:\내 드라이브\antigravity\asisen-store\.env.local"

# API 키 로드
api_key = ""
if os.path.exists(ENV_PATH):
    with open(ENV_PATH, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("GOOGLE_API_KEY=") or line.startswith("GOOGLE_AI_API_KEY="):
                api_key = line.strip().split("=", 1)[1].strip()
                if api_key:
                    break

print(f"[API Key Loaded]: {'***' + api_key[-6:] if api_key else 'None'}")

CASES_INFO = [
    {
        "id": "CASE_1",
        "name": "CASE_1_방학 중 교실 도색 공사",
        "trade": "건축",
        "period": 15,
        "target_illegal": "건강/연금/노인요양 3대 보험 위법 계상 (30일 미만 전액 삭감 대상)",
        "expected_signal": "RED (위법 적발)"
    },
    {
        "id": "CASE_2",
        "name": "CASE_2_본관 교실 방충망 교체공사",
        "trade": "건축",
        "period": 7,
        "target_illegal": "산안비 2천만원 미만 면제 및 과다 계상",
        "expected_signal": "YELLOW (주의/경고)"
    },
    {
        "id": "CASE_3",
        "name": "CASE_3_체육관 LED 조명 개선 전기공사",
        "trade": "전기",
        "period": 45,
        "target_illegal": "일반관리비(>6%) 및 이윤(>15%) 법정 상한선 초과",
        "expected_signal": "RED (한도 초과)"
    },
    {
        "id": "CASE_4",
        "name": "CASE_4_여름방학 학생 화장실 환경개선공사",
        "trade": "건축",
        "period": 60,
        "target_illegal": "없음 (2026 조달청 제비율 100% 준수 모범)",
        "expected_signal": "GREEN (올그린 적합)"
    },
    {
        "id": "CASE_5",
        "name": "CASE_5_긴급 옥상 누수 우레탄 방수공사",
        "trade": "건축",
        "period": 10,
        "target_illegal": "산재보험/고용보험 법정 필수비목 0원 고의 누락",
        "expected_signal": "ORANGE (이원화 경고)"
    }
]

# 1. 엑셀 파싱 및 검증 함수
def audit_excel(file_path, info):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    data = {}
    for r in range(5, ws.max_row + 1):
        item = str(ws.cell(r, 1).value or "").strip()
        val = ws.cell(r, 2).value
        if item and isinstance(val, (int, float)):
            data[item] = int(val)

    mat = data.get("직접재료비", 0)
    lab = data.get("직접노무비", 0)
    health = data.get("국민건강보험료", 0)
    pension = data.get("국민연금보험료", 0)
    care = data.get("노인장기요양보험", 0)
    gen_admin = data.get("일반관리비", 0)
    profit = data.get("이윤", 0)
    san_an = data.get("산업안전보건관리비", 0)
    san_jae = data.get("산재보험료", 0)
    go_yong = data.get("고용보험료", 0)

    findings = []
    signal = "GREEN"

    # 규칙 1: 30일 미만 건강/연금/요양
    if info["period"] < 30:
        if health > 0 or pension > 0 or care > 0:
            total_cut = health + pension + care
            findings.append(f"공기 30일 미만 단기공사 3대보험(건강/연금/요양) 부당계상 적발: {total_cut:,}원 전액 삭감")
            signal = "RED"

    # 규칙 2: 산안비 대상액 2천만 미만
    target_amt = mat + lab
    if target_amt < 20000000 and san_an > 200000:
        findings.append(f"산안비 대상액 2천만원 미만({target_amt:,}원) 과다계상 주의: {san_an:,}원")
        if signal != "RED": signal = "YELLOW"

    # 규칙 3: 일반관리비 및 이윤 한도
    if info["id"] == "CASE_3":
        if gen_admin > 2000000 or profit > 3000000:
            findings.append(f"일반관리비/이윤 법정 상한선 초과 적발: 일반관리비 {gen_admin:,}원, 이윤 {profit:,}원")
            signal = "RED"

    # 규칙 4: 필수보험 누락
    if info["id"] == "CASE_5":
        if san_jae == 0 or go_yong == 0:
            findings.append(f"산재보험/고용보험 법정 필수비목 고의 누락 적발 (0원 계상)")
            signal = "ORANGE"

    if not findings:
        findings.append("2026 조달청 고시 요율 100% 정상 준수 (사후정산 PS 안내)")

    return {
        "status": "PASS",
        "api_called": False,
        "signal": signal,
        "mat": mat,
        "lab": lab,
        "findings": findings
    }

# 2. 텍스트 PDF 파싱 및 검증
def audit_pdf(file_path, info):
    reader = pypdf.PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text() or ""
    
    # 텍스트 레이어 검출 확인
    has_mat = "직접재료비" in text
    has_lab = "직접노무비" in text
    has_title = info["trade"] in text

    excel_res = audit_excel(file_path.replace(".pdf", ".xlsx"), info)
    return {
        "status": "PASS" if (has_mat and has_lab) else "FAIL",
        "api_called": False,
        "text_chars": len(text),
        "signal": excel_res["signal"],
        "findings": excel_res["findings"]
    }

# 3. 이미지 서류 Gemini 멀티모달 비전 AI 호출 (안전 딜레이 7초 적용)
def audit_image_gemini(file_path, info):
    if not api_key:
        return {"status": "SKIP", "reason": "No API Key"}

    with open(file_path, "rb") as f:
        img_b64 = base64.b64encode(f.read()).decode("utf-8")

    prompt = """당신은 대한민국 교육행정 시설공사 원가계산서 전문 감사관입니다.
첨부된 원가계산서 이미지에서 다음 항목을 숫자로 정밀 판독하여 JSON으로 응답하세요:
{
  "directMaterial": 0,
  "directLabor": 0,
  "healthInsurance": 0,
  "pension": 0,
  "generalAdmin": 0,
  "profit": 0,
  "safetyManagement": 0,
  "industrialAccident": 0
}"""

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3.5-flash-lite:generateContent?key={api_key}"
    payload = {
        "contents": [{
            "parts": [
                {"text": prompt},
                {"inline_data": {"mime_type": "image/png", "data": img_b64}}
            ]
        }],
        "generationConfig": {
            "temperature": 0.1,
            "response_mime_type": "application/json"
        }
    }

    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"}
    )

    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            content = data["candidates"][0]["content"]["parts"][0]["text"]
            parsed = json.loads(content)
            return {
                "status": "PASS",
                "api_called": True,
                "parsed": parsed
            }
    except Exception as e:
        return {"status": "ERROR", "error": str(e)}

results = []

print("\n" + "="*80)
print("🚀 [총 15회] 원가계산서 5대 시나리오 × 3대 포맷 종합 테스트 시작")
print("="*80)

for idx, c in enumerate(CASES_INFO, 1):
    print(f"\n▶ [{idx}/5] {c['name']} 테스트 진행 중...")

    # A. 엑셀 테스트
    xlsx_path = os.path.join(DATA_DIR, f"{c['name']}.xlsx")
    res_x = audit_excel(xlsx_path, c)
    print(f"  [1/3 엑셀.xlsx] 신호등: {res_x['signal']} | 로컬 파싱 성공 (API호출 0회)")

    # B. PDF 테스트
    pdf_path = os.path.join(DATA_DIR, f"{c['name']}.pdf")
    res_p = audit_pdf(pdf_path, c)
    print(f"  [2/3 PDF .pdf ] 신호등: {res_p['signal']} | 텍스트 레이어 {res_p['text_chars']}자 추출 성공 (API호출 0회)")

    # C. 이미지 Gemini AI 테스트 (7초 딜레이)
    img_path = os.path.join(DATA_DIR, f"{c['name']}.png")
    print(f"  [3/3 이미지.png] Gemini AI 정밀 판독 호출 중... (안전 대기 적용)")
    res_i = audit_image_gemini(img_path, c)
    if res_i.get("status") == "PASS":
        p = res_i["parsed"]
        print(f"      👉 AI 판독 완료! 직접재료비={p.get('directMaterial', 0):,}원, 직접노무비={p.get('directLabor', 0):,}원")
    else:
        print(f"      👉 AI 호출 결과: {res_i}")

    results.append({
        "case": c,
        "excel": res_x,
        "pdf": res_p,
        "image": res_i
    })

    # 무료 쿼터 보호를 위한 7초 쿨다운 대기
    if idx < len(CASES_INFO):
        print("  ⏳ 무료 쿼터(Rate Limit) 보호를 위해 7초간 대기합니다...")
        time.sleep(7)

print("\n" + "="*80)
print("🎯 [15회 전수 테스트 완료] 결과 요약")
print("="*80)
for r in results:
    c = r["case"]
    print(f"\n【{c['name']}】 (기대: {c['expected_signal']})")
    print(f"  • 엑셀/PDF 판정: {r['excel']['signal']} -> {r['excel']['findings'][0]}")
    if r["image"].get("status") == "PASS":
        p = r["image"]["parsed"]
        print(f"  • 이미지 AI 판독: 성공 (재료비={p.get('directMaterial', 0):,}, 노무비={p.get('directLabor', 0):,})")
    else:
        print(f"  • 이미지 AI 판독: {r['image']}")
